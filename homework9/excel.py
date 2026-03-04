from openpyxl import Workbook, load_workbook
# --- Zapis do pliku Excel --
wb = Workbook()
ws = wb.active # Aktywny arkusz
ws.title = "Użytkownicy"
ws.append(["Imię", "Wiek"]) # Dodaj nagłówek
ws.append(["Anna", 25])
ws.append(["Piotr", 32])
wb.save("uzytkownicy.xlsx")
# --- Odczyt z pliku Excel --
wb_odczyt = load_workbook("uzytkownicy.xlsx")
ws_odczyt = wb_odczyt.active
for wiersz in ws_odczyt.iter_rows(values_only=True):
    print(wiersz)